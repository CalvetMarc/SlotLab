import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import json


def load_game_tables(base_path):
    """
    Automatically loads all game tables from a given game directory.
    The function assumes the folder contains:
      - mathsTables.json → defines what and how to load
      - slotMaths.xlsx   → contains the actual data
    """

    base_path = Path(base_path)
    config_json_path = base_path / "mathsTables.json"
    excel_file = base_path / "slotMaths.xlsx"

    # --- Validate files ---
    if not config_json_path.exists():
        raise FileNotFoundError(f"❌ mathsTables.json not found in {base_path}")
    if not excel_file.exists():
        raise FileNotFoundError(f"❌ slotMaths.xlsx not found in {base_path}")

    print(f"📄 Reading config from: {config_json_path}")
    print(f"📊 Reading Excel data from: {excel_file}\n")

    # --- Load configuration JSON ---
    with open(config_json_path, "r", encoding="utf-8") as f:
        config = json.load(f)

    # Build a list of all table configurations
    table_configs = []
    for group, tables in config.items():
        for t in tables:
            if isinstance(t, str):
                # Backward compatibility: simple string = table name
                table_configs.append({"name": t, "orientation": "records"})
            else:
                table_configs.append(t)

    # --- Load Excel workbook ---
    wb = load_workbook(excel_file, data_only=True)
    result = {}

    # --- Iterate through all configured tables ---
    for tconf in table_configs:
        table_name = tconf["name"]
        lname = table_name.lower()
        orientation = tconf.get("orientation", "records")
        ignore_columns = tconf.get("ignore_columns", [])
        multiplier = tconf.get("multiply_values_by", None)

        df = None
        found_table = False

        # 1️⃣ Try to load as a defined Excel Table (named range)
        for ws in wb.worksheets:
            if table_name in ws.tables:
                table_obj = ws.tables[table_name]
                ref = table_obj.ref
                start_cell, end_cell = ref.split(":")
                start_col = ''.join([c for c in start_cell if c.isalpha()])
                start_row = int(''.join([c for c in start_cell if c.isdigit()]))
                end_col = ''.join([c for c in end_cell if c.isalpha()])
                end_row = int(''.join([c for c in end_cell if c.isdigit()]))

                df = pd.read_excel(
                    excel_file,
                    sheet_name=ws.title,
                    header=None,
                    usecols=f"{start_col}:{end_col}",
                    skiprows=start_row - 1,
                    nrows=end_row - start_row + 1
                )

                df.columns = df.iloc[0]
                df = df[1:].reset_index(drop=True)
                found_table = True
                break

        # 2️⃣ If no named table found, try to read by sheet name
        if not found_table:
            if table_name in wb.sheetnames:
                print(f"📄 Loading sheet '{table_name}' directly (no named Excel Table found).")
                df = pd.read_excel(excel_file, sheet_name=table_name)
            else:
                print(f"⚠️ Table or sheet '{table_name}' not found in workbook. Skipping.")
                continue

        # --- Post-processing ---
        if ignore_columns:
            df = df[[c for c in df.columns if c not in ignore_columns]]

        if multiplier:
            df = df.map(lambda x: x * multiplier if isinstance(x, (int, float)) else x)

        # Convert to dict depending on orientation
        if orientation == "list":
            processed = {col: df[col].dropna().tolist() for col in df.columns}
        else:  # default: records
            processed = df.to_dict(orient="records")

        result[lname] = processed
        print(f"✅ Loaded and processed table: {table_name} ({orientation})")

    print(f"\n🎯 Total tables processed: {list(result.keys())}")
    return result
